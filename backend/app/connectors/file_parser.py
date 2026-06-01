"""Excel / CSV 文件解析器"""

from __future__ import annotations

import csv
import os
import sys
from pathlib import Path
from app.connectors.base import DataSourceConnector
from app.models import TableInfo, ColumnInfo

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent.parent
UPLOAD_DIR = BASE_DIR / "uploads"


class FileParserConnector(DataSourceConnector):
    """解析上传的 Excel/CSV 文件，提取 Sheet/列结构"""

    def __init__(self):
        self._tables: list[TableInfo] = []
        self._columns: dict[str, list[ColumnInfo]] = {}
        self._file_path: str = ""
        self._file_type: str = ""

    async def connect(self, config: dict) -> bool:
        self._file_path = config["file_path"]
        self._file_type = config.get("file_type", "")
        if not self._file_type:
            ext = os.path.splitext(self._file_path)[1].lower()
            self._file_type = ext

        if self._file_type in (".xlsx", ".xls"):
            return await self._parse_excel()
        elif self._file_type == ".csv":
            return await self._parse_csv()
        else:
            raise ValueError(f"不支持的文件格式: {self._file_type}")

    async def disconnect(self):
        self._tables = []
        self._columns = {}
        self._file_path = ""
        self._file_type = ""

    async def test_connection(self, config: dict) -> bool:
        path = config.get("file_path", "")
        return os.path.isfile(path)

    async def get_tables(self) -> list[TableInfo]:
        return self._tables

    async def get_columns(self, table: str, schema: str = "") -> list[ColumnInfo]:
        return self._columns.get(table, [])

    async def preview_data(self, table: str, schema: str = "", limit: int = 10) -> list[dict]:
        if self._file_type == ".csv":
            return await self._preview_csv(limit)
        elif self._file_type in (".xlsx", ".xls"):
            return await self._preview_excel(table, limit)
        return []

    # ---- Excel 解析 ----

    async def _parse_excel(self) -> bool:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
            return await self._parse_openpyxl(wb)
        except Exception:
            try:
                import xlrd
                wb = xlrd.open_workbook(self._file_path)
                return await self._parse_xlrd(wb)
            except Exception:
                raise RuntimeError("无法打开 Excel 文件，请安装 openpyxl 或 xlrd")

    async def _parse_openpyxl(self, wb) -> bool:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(min_row=1, max_row=min(ws.max_row, 21), values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            if not header or all(h is None or str(h).strip() == "" for h in header):
                continue

            raw_header = [str(h).strip() if h is not None else f"Column{i}" for i, h in enumerate(header)]
            header = []
            seen = set()
            for h in raw_header:
                if h in seen:
                    count = 1
                    while f"{h}_{count}" in seen:
                        count += 1
                    h = f"{h}_{count}"
                seen.add(h)
                header.append(h)

            col_types = []
            row_count = 0
            for row in rows_iter:
                if row and any(v is not None for v in row):
                    row_count += 1
                    if row_count <= 20:
                        for i, val in enumerate(row):
                            if i < len(col_types):
                                if val is not None:
                                    col_types[i] = self._merge_type(col_types[i], type(val).__name__)
                            else:
                                col_types.append(type(val).__name__ if val is not None else "str")

            while len(col_types) < len(header):
                col_types.append("str")

            type_map = {"int": "INTEGER", "float": "FLOAT", "str": "VARCHAR", "bool": "BOOLEAN"}
            columns = [
                ColumnInfo(name=h, data_type=type_map.get(col_types[i], "VARCHAR"))
                for i, h in enumerate(header)
            ]
            self._tables.append(TableInfo(name=sheet_name, row_count=row_count))
            self._columns[sheet_name] = columns

        wb.close()
        return True

    async def _parse_xlrd(self, wb) -> bool:
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            if ws.nrows < 1:
                continue
            raw_header = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            header = []
            seen = set()
            for h in raw_header:
                if h in seen:
                    count = 1
                    while f"{h}_{count}" in seen:
                        count += 1
                    h = f"{h}_{count}"
                seen.add(h)
                header.append(h)
            preview_rows = min(ws.nrows, 21)
            col_types = []
            for r in range(1, preview_rows):
                for c in range(ws.ncols):
                    val = ws.cell_value(r, c)
                    if val != "" and val is not None:
                        if c < len(col_types):
                            col_types[c] = self._merge_type(col_types[c], type(val).__name__)
                        else:
                            col_types.append(type(val).__name__)
            while len(col_types) < len(header):
                col_types.append("str")
            type_map = {"int": "INTEGER", "float": "FLOAT", "str": "VARCHAR", "bool": "BOOLEAN"}
            columns = [
                ColumnInfo(name=h, data_type=type_map.get(col_types[i], "VARCHAR"))
                for i, h in enumerate(header)
            ]
            self._tables.append(TableInfo(name=ws.name, row_count=ws.nrows - 1))
            self._columns[ws.name] = columns
        return True

    async def _preview_excel(self, sheet: str, limit: int) -> list[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        except Exception:
            return []
        if sheet not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet]
        rows_iter = ws.iter_rows(min_row=1, max_row=min(ws.max_row, limit + 1), values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            wb.close()
            return []
        raw_header = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(header)]
        header = []
        seen = set()
        for h in raw_header:
            if h in seen:
                count = 1
                while f"{h}_{count}" in seen:
                    count += 1
                h = f"{h}_{count}"
            seen.add(h)
            header.append(h)
        result = []
        for row in rows_iter:
            result.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
        wb.close()
        return result

    # ---- CSV 解析 ----

    async def _parse_csv(self) -> bool:
        try:
            import chardet
        except ImportError:
            chardet = None

        encoding = "utf-8"
        with open(self._file_path, "rb") as f:
            raw = f.read(4096)
            if chardet:
                detected = chardet.detect(raw)
                if detected and detected.get("encoding"):
                    encoding = detected["encoding"]

        with open(self._file_path, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = "excel"
            reader = csv.reader(f, dialect)
            try:
                header = next(reader)
            except StopIteration:
                return True
            raw_header = [h.strip() for h in header if h.strip()]
            if not raw_header:
                return True

            header = []
            seen = set()
            for h in raw_header:
                if h in seen:
                    count = 1
                    while f"{h}_{count}" in seen:
                        count += 1
                    h = f"{h}_{count}"
                seen.add(h)
                header.append(h)

            col_types = []
            row_count = 0
            for row in reader:
                if row:
                    row_count += 1
                    if row_count <= 20:
                        for i, val in enumerate(row):
                            inferred = self._infer_csv_type(val)
                            if i < len(col_types):
                                col_types[i] = self._merge_type(col_types[i], inferred)
                            else:
                                col_types.append(inferred)

        while len(col_types) < len(header):
            col_types.append("VARCHAR")

        file_name = os.path.basename(self._file_path)
        columns = [
            ColumnInfo(name=h, data_type=col_types[i])
            for i, h in enumerate(header)
        ]
        self._tables.append(TableInfo(name=file_name, row_count=row_count))
        self._columns[file_name] = columns
        return True

    async def _preview_csv(self, limit: int) -> list[dict]:
        try:
            import chardet
        except ImportError:
            chardet = None

        encoding = "utf-8"
        with open(self._file_path, "rb") as f:
            raw = f.read(4096)
            if chardet:
                detected = chardet.detect(raw)
                if detected and detected.get("encoding"):
                    encoding = detected["encoding"]

        with open(self._file_path, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = "excel"
            reader = csv.reader(f, dialect=dialect)
            try:
                header = next(reader)
            except StopIteration:
                return []
            
            raw_header = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(header)]
            header = []
            seen = set()
            for h in raw_header:
                if h in seen:
                    count = 1
                    while f"{h}_{count}" in seen:
                        count += 1
                    h = f"{h}_{count}"
                seen.add(h)
                header.append(h)
            
            rows = []
            for i, row in enumerate(reader):
                if i >= limit:
                    break
                rows.append({header[j]: row[j] if j < len(row) else None for j in range(len(header))})
            return rows

    def _infer_csv_type(self, val: str) -> str:
        if val is None or val.strip() == "":
            return "VARCHAR"
        val = val.strip()
        try:
            int(val)
            return "INTEGER"
        except ValueError:
            pass
        try:
            float(val)
            return "FLOAT"
        except ValueError:
            pass
        return "VARCHAR"

    def _merge_type(self, old: str, new: str) -> str:
        if old == new:
            return old
        if old == "VARCHAR" or new == "VARCHAR":
            return "VARCHAR"
        if "float" in (old.lower(), new.lower()):
            return "FLOAT"
        return "VARCHAR"
